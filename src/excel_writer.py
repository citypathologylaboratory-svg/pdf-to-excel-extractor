"""Excel file writer module for exporting data to Excel format."""

from pathlib import Path
from typing import Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from loguru import logger


class ExcelWriter:
    """Handles Excel file creation and formatting."""

    def __init__(self, auto_format: bool = True):
        """Initialize the Excel writer.
        
        Args:
            auto_format: Whether to apply automatic formatting
        """
        self.auto_format = auto_format

    def write_dataframe(self, df: pd.DataFrame, output_path: str,
                       sheet_name: str = 'Data') -> None:
        """Write a pandas DataFrame to an Excel file.
        
        Args:
            df: DataFrame to write
            output_path: Output file path
            sheet_name: Name of the Excel sheet
        """
        try:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Create workbook and sheet
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Write headers
            for col_idx, column in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=column)
                if self.auto_format:
                    self._style_header(cell)
            
            # Write data
            for row_idx, row in enumerate(df.values, 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if self.auto_format:
                        self._style_cell(cell)
            
            # Adjust column widths
            if self.auto_format:
                self._auto_adjust_columns(ws, df)
            
            # Save workbook
            wb.save(output_path)
            logger.success(f"Excel file created: {output_path}")
            
        except Exception as e:
            logger.error(f"Error writing Excel file: {str(e)}")
            raise

    def _style_header(self, cell) -> None:
        """Apply header styling to a cell."""
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="366092", end_color="366092",
                               fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                  wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _style_cell(self, cell) -> None:
        """Apply standard cell styling."""
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                  wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _auto_adjust_columns(self, worksheet, df: pd.DataFrame) -> None:
        """Auto-adjust column widths based on content."""
        for col_idx, column in enumerate(df.columns, 1):
            max_length = len(str(column)) + 2
            for value in df[column]:
                if value is not None:
                    max_length = max(max_length, len(str(value)) + 2)
            
            worksheet.column_dimensions[chr(64 + col_idx)].width = min(
                max_length, 50
            )

    def write_multiple_sheets(self, data_dict: dict, output_path: str) -> None:
        """Write multiple DataFrames to different sheets in one Excel file.
        
        Args:
            data_dict: Dictionary with sheet names as keys and DataFrames as values
            output_path: Output file path
        """
        try:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            for sheet_name, df in data_dict.items():
                ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit
                
                # Write headers
                for col_idx, column in enumerate(df.columns, 1):
                    cell = ws.cell(row=1, column=col_idx, value=column)
                    if self.auto_format:
                        self._style_header(cell)
                
                # Write data
                for row_idx, row in enumerate(df.values, 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        if self.auto_format:
                            self._style_cell(cell)
                
                if self.auto_format:
                    self._auto_adjust_columns(ws, df)
            
            wb.save(output_path)
            logger.success(f"Multi-sheet Excel file created: {output_path}")
            
        except Exception as e:
            logger.error(f"Error writing multi-sheet Excel file: {str(e)}")
            raise
